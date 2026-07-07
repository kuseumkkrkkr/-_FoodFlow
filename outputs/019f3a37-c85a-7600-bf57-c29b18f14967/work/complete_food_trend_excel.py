from __future__ import annotations

from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment


ROOT = Path(__file__).resolve().parents[3]
OUTPUT_DIR = ROOT / "outputs" / "019f3a37-c85a-7600-bf57-c29b18f14967"
SOURCE = next(ROOT.glob("1.A02*.xlsx"))
TARGET = OUTPUT_DIR / "1.A02_SCOUT_THE_100_식품트렌드매핑_Tool_완성_10트렌드.xlsx"


ROWS = [
    {
        "no": 1,
        "keyword": "기능성·개인맞춤 식품",
        "meaning": "단백질, 식이섬유, 장건강, 에너지·멘탈 관리처럼 효능이 명확한 제품을 개인 목적에 맞춰 선택·기획하는 흐름",
        "background": "2026 식품 트렌드에서 단백질, 장건강, 기능성 음료와 개인화 영양이 핵심으로 제시됨. 브랜드는 기능성 원료와 제형에 맞는 빠른 제품화가 필요함",
        "consumer": "소비자는 단순한 맛보다 건강 목적, 성분 투명성, 섭취 상황별 편의성을 함께 요구함",
        "cases": "고단백 스낵·음료, 프로/프리바이오틱스 제품, 기능성 RTD 음료, 개인화 영양 앱",
        "connection": "입력한 제품 컨셉의 기능성 원료, 제형, 인증 조건을 분석해 생산 가능한 OEM/ODM 공장을 추천할 수 있음",
        "category": "음료, 단백질바, 건강간식, 발효식품, 분말/스틱",
        "memo": "공장 DB에는 취급 원료, MOQ, 인증, 소량 샘플 가능 여부를 태그화해야 함",
        "question": "기능성 식품 창업자가 공장 탐색에서 가장 오래 걸리는 조건은 무엇인가?",
        "source": "Obsidian_Food_OEM_ODM/trends/Functional_Personalized_Food.md",
    },
    {
        "no": 2,
        "keyword": "AI·투명 공급망 제조",
        "meaning": "AI와 데이터로 원료, 공정, 인증, 생산 가능 수량, 납기 정보를 연결해 식품 개발·발주를 빠르게 만드는 흐름",
        "background": "AI, 바이오프로세싱, 디지털 추적 기술이 식품 설계·공급망 투명성·안전성에 활용되고, 클린라벨과 책임소싱 요구가 기본 조건으로 확대됨",
        "consumer": "소규모 브랜드도 원산지, 제조공정, 클린라벨, 지속가능성을 설명할 수 있는 제조 파트너를 원함",
        "cases": "AI 레시피 개발, 스마트 패키징/QR, 추적 가능한 원료 소싱, 소량 생산 D2C 제조 플랫폼",
        "connection": "B2B 대량 발주와 B2C 소규모 제작/공동구매를 분리해 MOQ, 위치, 설비 특성별 최적 공장 매칭이 가능함",
        "category": "소스/드레싱, HMR, 간편식, 베이커리, 펫푸드, 로컬 특산가공",
        "memo": "핵심은 공장 역량 데이터 표준화와 견적 자동화. 단순 중개보다 매칭 정확도와 신뢰 검증이 차별점",
        "question": "공장과 발주자가 모두 제공할 수 있는 필수 데이터 항목은 무엇인가?",
        "source": "Obsidian_Food_OEM_ODM/trends/AI_Transparent_Supply_Chain.md",
    },
    {
        "no": 3,
        "keyword": "단백질·장건강",
        "meaning": "고단백, 식이섬유, 발효, 프로/프리바이오틱스 제품이 건강 관리와 포만감, 소화 개선 니즈에 맞춰 확대되는 흐름",
        "background": "기능성·개인맞춤 식품 흐름 안에서 단백질, 장건강, 발효 기반 제품이 반복적으로 강조됨. 원료 안정성, 균주 관리, 가열·보관 조건이 제조 난도를 높임",
        "consumer": "소비자는 포만감, 소화, 면역, 에너지 관리처럼 일상 건강 문제를 해결하는 스낵·음료·발효식품을 찾음",
        "cases": "단백질바, 그래놀라바, 저당 고식이섬유 바, 프로틴 쿠키, 프로바이오틱 음료, 발효 간식",
        "connection": "단백질·장건강 제품은 원료 안정성, 균주, 가열 공정, 냉장/냉동 유통 조건이 달라 전문 공장 매칭이 필요함",
        "category": "단백질바, 건강간식, 기능성 음료, 발효식품, 분말/스틱",
        "memo": "반복 발주와 소량 테스트에 적합하므로 B2C 제작부와 B2B PB/OEM 전환을 함께 검증하기 좋음",
        "question": "단백질·장건강 제품에서 발주자가 가장 먼저 확인해야 하는 공장 조건은 원료, 공정, 인증, MOQ 중 무엇인가?",
        "source": "Obsidian_Food_OEM_ODM/trends/Protein_Gut_Health.md",
    },
    {
        "no": 4,
        "keyword": "클린라벨·책임소싱",
        "meaning": "소비자가 원재료, 원산지, 제조 과정, 지속가능성 정보를 더 투명하게 요구하는 흐름",
        "background": "식품 구매에서 원재료 목록, 인증, 영양정보, 알레르기, 제조방식, 조달 방식 같은 정보 공개가 신뢰 판단 기준으로 확대됨",
        "consumer": "소비자는 착한 기업 이미지보다 실제 성분, 원산지, 인증, 알레르기 관리와 같은 확인 가능한 정보를 보고 구매를 결정함",
        "cases": "클린라벨 스낵, 원산지 추적 원료, 유기농·비건·할랄 인증 제품, QR 기반 제조정보 공개, 친환경 포장",
        "connection": "공장별 인증, 원료 추적, 알레르기 관리, 첨가물 정책을 데이터화해 비교할 수 있음",
        "category": "건강간식, 소스/드레싱, 기능성 음료, 베이커리, 로컬 특산가공",
        "memo": "플랫폼은 생산 가능 여부뿐 아니라 증빙 문서 유무와 소비자 공개 가능 범위를 함께 관리해야 함",
        "question": "발주자가 소비자에게 공개해야 하는 원재료, 원산지, 알레르기, 인증, 제조방식 중 가장 부담되는 항목은 무엇인가?",
        "source": "Obsidian_Food_OEM_ODM/trends/Clean_Label_Responsible_Sourcing.md",
    },
    {
        "no": 5,
        "keyword": "편의성·가성비",
        "meaning": "가격 부담 속에서도 품질, 건강성, 편의성을 함께 요구하며 작게 테스트하고 빠르게 구매하는 흐름",
        "background": "소비자는 접근성과 가격을 중시하고, 초기 식품 브랜드는 대량 생산 전 작은 수량으로 시장성을 확인해야 하는 부담이 커짐",
        "consumer": "크리에이터, 로컬 브랜드, 커뮤니티 운영자, 예비 창업자는 낮은 초기 비용과 빠른 샘플 제작, 사전 수요 확인을 원함",
        "cases": "소량 제작 패키지, 공동구매형 생산, 사전예약 테스트 판매, 샘플 키트, 크리에이터 한정판 식품",
        "connection": "B2C 소규모 제작부와 공동구매형 생산을 통해 대량 생산 전 시장 검증을 가능하게 함",
        "category": "건강간식, 단백질바, 분말/스틱, 베이커리, 로컬 가공식품",
        "memo": "표준 옵션, 예상 단가, 최소 수량, 제작 기간을 먼저 보여주면 초보 발주자의 진입 장벽을 낮출 수 있음",
        "question": "소량 제작 사용자는 공장 추천 전 예상 단가, 최소 수량, 제작 기간, 샘플 가능 여부 중 무엇을 먼저 알고 싶어 하는가?",
        "source": "Obsidian_Food_OEM_ODM/trends/Convenience_Value.md",
    },
    {
        "no": 6,
        "keyword": "목적형 기능성 음료",
        "meaning": "수분 보충, 장건강, 에너지, 집중, 릴랙스, 뷰티처럼 명확한 섭취 목적을 가진 음료가 확대되는 흐름",
        "background": "2026 식품 트렌드에서 기능성 음료와 목적 기반 음료가 주요 카테고리로 부상함. 소비자는 단순 갈증 해소보다 효능, 성분, 섭취 상황을 함께 봄",
        "consumer": "소비자는 운동 전후, 업무 집중, 수면 전, 장건강 관리 등 상황별로 마실 이유가 분명한 음료를 선택하려 함",
        "cases": "단백질 RTD, 전해질 음료, 프로바이오틱 음료, 카페인·무카페인 에너지 음료, 릴랙스·수면 보조 음료",
        "connection": "음료는 제형, 살균 방식, 충진 설비, 기능성 원료 안정성, 유통 온도 조건이 달라 공장 매칭 난도가 높음",
        "category": "RTD 음료, 분말 스틱, 기능성 워터, 발효 음료, 프로틴 음료",
        "memo": "공장 DB에 액상/분말, 병·캔·파우치 충진, 살균 방식, 기능성 원료 대응 여부를 세분화해야 함",
        "question": "기능성 음료 창업자는 제조처 선택에서 제형, 충진 방식, 기능성 원료 안정성, MOQ 중 무엇을 가장 먼저 확인하는가?",
        "source": "Innova Market Insights Top Food and Beverage Trends 2026; https://www.innovamarketinsights.com/trends/top-food-and-beverage-trends-2026/",
    },
    {
        "no": 7,
        "keyword": "식물성 2.0·대체단백",
        "meaning": "모방육 중심의 식물성 식품을 넘어 콩, 곡물, 버섯, 해조류, 발효 단백질처럼 맛과 영양, 지속가능성을 함께 설계하는 흐름",
        "background": "대체단백은 단순한 고기 대체에서 벗어나 원료 다양화, 발효·효소 기술, 영양 균형, 클린라벨 요구와 결합하는 방향으로 진화함",
        "consumer": "소비자는 식물성이라서가 아니라 맛, 단백질 함량, 가격, 원재료 신뢰도, 환경적 의미가 납득될 때 구매함",
        "cases": "식물성 단백질바, 버섯 기반 스낵, 해조류 원료 제품, 콩·완두 단백 음료, 발효 단백 원료",
        "connection": "대체단백 제품은 원료 취급, 배합, 식감 구현, 알레르기 관리, 표시 기준이 중요해 전문 제조처 필터링이 필요함",
        "category": "단백질바, 스낵, 식물성 음료, HMR, 베이커리, 분말/스틱",
        "memo": "단백질 원료별 알레르기, 맛 마스킹, 식감 구현 설비, 비건 인증 대응 가능 여부를 매칭 기준에 포함해야 함",
        "question": "식물성 제품 발주자는 맛, 단백질 함량, 원료 원산지, 비건 인증, 단가 중 어떤 조건을 우선순위로 보는가?",
        "source": "Kerry Health and Nutrition Institute 2026 megatrends; https://khni.kerry.com/key-health-and-nutrition-trends/",
    },
    {
        "no": 8,
        "keyword": "멘탈 밸런스·수면 케어",
        "meaning": "스트레스, 집중, 기분, 수면 같은 정신적 컨디션 관리를 식품·음료 섭취 루틴과 연결하는 흐름",
        "background": "웰니스 소비가 신체 건강에서 감정·수면·집중 관리로 확장되며, 기능성 원료와 섭취 경험을 결합한 제품 기획이 늘고 있음",
        "consumer": "소비자는 카페인 대체, 숙면 루틴, 업무 집중, 긴장 완화처럼 일상 상황에 맞는 저부담 웰니스 제품을 찾음",
        "cases": "릴랙스 티, 무알코올 기능성 음료, L-테아닌·마그네슘 함유 제품, 수면 루틴 젤리, 집중력 보조 스낵",
        "connection": "멘탈·수면 케어 제품은 원료 표시, 기능성 표현 가능 범위, 맛 설계, 제형 안정성을 함께 검토해야 함",
        "category": "차/티백, 기능성 음료, 젤리, 건강간식, 분말/스틱",
        "memo": "플랫폼은 과장 표시 리스크를 줄이기 위해 원료, 표시 문구, 인증·시험자료 보유 여부를 공장 데이터와 연결해야 함",
        "question": "멘탈·수면 케어 제품에서 소비자가 신뢰하는 근거는 원료명, 함량, 시험자료, 리뷰, 브랜드 설명 중 무엇인가?",
        "source": "Innova Market Insights Top Food and Beverage Trends 2026; https://www.innovamarketinsights.com/trends/top-food-and-beverage-trends-2026/",
    },
    {
        "no": 9,
        "keyword": "로컬·전통 발효 재해석",
        "meaning": "지역 원료, 전통 조리법, 발효 기술을 현대적인 맛·패키지·건강 가치로 재해석하는 흐름",
        "background": "소비자는 새로운 맛을 원하면서도 진정성, 원산지, 스토리, 전통성에 반응함. 발효와 로컬 소싱은 건강·지속가능성 메시지와도 연결됨",
        "consumer": "소비자는 흔한 건강식보다 지역성, 제조 스토리, 전통 방식, 믿을 수 있는 원료가 있는 제품에 더 높은 의미를 부여함",
        "cases": "로컬 과일 발효 음료, 전통 장류 소스, 김치·피클류, 지역 특산 스낵, 발효 디저트",
        "connection": "로컬·발효 제품은 원료 수급, 발효 설비, 냉장 유통, 소량 생산, 지역 공장 네트워크가 중요해 위치 기반 매칭과 잘 맞음",
        "category": "발효식품, 소스/드레싱, 음료, 로컬 특산가공, HMR",
        "memo": "공장 DB에 지역, 원료 수급처, 발효 가능 설비, 냉장/냉동 유통, 관광·선물용 패키징 대응 여부를 추가해야 함",
        "question": "로컬 원료 기반 식품은 지역 공장 생산, 원료 증빙, 소량 생산, 스토리텔링 중 무엇이 구매 설득에 가장 큰 영향을 주는가?",
        "source": "Innova Market Insights Top Food and Beverage Trends 2026; https://www.innovamarketinsights.com/trends/top-food-and-beverage-trends-2026/",
    },
    {
        "no": 10,
        "keyword": "텍스처·작은 사치 스낵",
        "meaning": "큰 지출 대신 바삭함, 쫀득함, 층감, 미니 사이즈, 프리미엄 패키지처럼 감각적 만족을 주는 간식 소비가 늘어나는 흐름",
        "background": "물가 부담이 이어지면서 소비자는 합리적인 가격 안에서 기분 전환과 재미를 주는 식품을 선택함. 맛뿐 아니라 식감과 사용 장면이 차별화 요소가 됨",
        "consumer": "소비자는 큰 용량보다 부담 없는 가격, 공유하기 좋은 크기, 사진에 잘 보이는 패키지, 새로운 식감 경험을 선호함",
        "cases": "미니 디저트, 프리미엄 한입 스낵, 층감 있는 바·쿠키, 모찌·젤리 식감 제품, 한정판 패키지 간식",
        "connection": "스낵·디저트는 배합, 성형, 굽기/건조, 코팅, 개별 포장 설비에 따라 생산 가능 공장이 달라져 조건 매칭이 중요함",
        "category": "건강간식, 디저트, 베이커리, 단백질바, 젤리, 키즈 간식",
        "memo": "B2C 제작부에서 샘플 키트, 한정판 소량 생산, 크리에이터 공동구매 상품으로 검증하기 좋음",
        "question": "소비자는 작은 사치 스낵을 고를 때 가격, 식감, 건강성, 패키지, 한정판 여부 중 무엇에 가장 반응하는가?",
        "source": "Innova Market Insights Top Food and Beverage Trends 2026; https://www.innovamarketinsights.com/trends/top-food-and-beverage-trends-2026/",
    },
]


def write_row(ws, row_idx: int, item: dict[str, str | int]) -> None:
    values = [
        item["no"],
        item["keyword"],
        item["meaning"],
        item["background"],
        item["consumer"],
        item["cases"],
        item["connection"],
        item["category"],
        item["memo"],
        item["question"],
    ]
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row_idx, col_idx)
        cell.value = value
        cell.alignment = copy(cell.alignment)
        cell.alignment = Alignment(
            horizontal="center" if col_idx == 1 else "left",
            vertical="center",
            wrap_text=True,
        )
    ws.cell(row_idx, 2).comment = Comment(f"Source: {item['source']}", "User")


def copy_row_style(ws, source_row: int, target_row: int, max_col: int) -> None:
    for col_idx in range(1, max_col + 1):
        source = ws.cell(source_row, col_idx)
        target = ws.cell(target_row, col_idx)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.protection:
            target.protection = copy(source.protection)
        if source.alignment:
            target.alignment = copy(source.alignment)


def write_merged(ws, address: str, value: str) -> None:
    cell = ws[addr_top_left(address)]
    cell.value = value
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)


def addr_top_left(address: str) -> str:
    return address.split(":")[0]


def set_font_size(cell, size: int) -> None:
    font = copy(cell.font)
    font.sz = size
    cell.font = font


def main() -> None:
    wb = load_workbook(SOURCE)
    guide = wb["작성가이드"]
    trend = wb["트렌드매핑_작성표"]
    card = wb["1page_요약카드"]

    guide.print_area = "A1:H24"
    guide.page_setup.orientation = "landscape"
    guide.page_setup.fitToWidth = 1
    guide.page_setup.fitToHeight = 1
    guide.sheet_properties.pageSetUpPr.fitToPage = True
    guide.page_margins.left = 0.25
    guide.page_margins.right = 0.25
    guide.page_margins.top = 0.4
    guide.page_margins.bottom = 0.4

    for row_idx in range(6, 16):
        copy_row_style(trend, 6, row_idx, 10)
        trend.row_dimensions[row_idx].height = 74

    for offset, item in enumerate(ROWS):
        write_row(trend, 6 + offset, item)

    for row_idx in range(16, 18):
        for col_idx in range(2, 11):
            trend.cell(row_idx, col_idx).value = None

    trend.column_dimensions["D"].width = 24
    trend.column_dimensions["J"].width = 26
    for col in "BCDEFGHIJ":
        trend.column_dimensions[col].bestFit = False
    trend.freeze_panes = "A6"
    trend.print_area = "A1:J22"
    trend.page_setup.orientation = "landscape"
    trend.page_setup.fitToWidth = 1
    trend.page_setup.fitToHeight = 1
    trend.sheet_properties.pageSetUpPr.fitToPage = True
    trend.page_margins.left = 0.25
    trend.page_margins.right = 0.25
    trend.page_margins.top = 0.4
    trend.page_margins.bottom = 0.4

    card["A4"] = "AI 기반 식품 OEM/ODM 공장 매칭 플랫폼"
    card["C4"] = "AI 기반 식품 OEM/ODM 공장 매칭 플랫폼"
    card["G4"] = "=TODAY()"

    write_merged(
        card,
        "A7:B8",
        "AI·투명 공급망 제조, 기능성·개인맞춤 식품, 단백질·장건강, 클린라벨·책임소싱, 편의성·가성비",
    )
    write_merged(
        card,
        "C7:D8",
        "식품 트렌드는 제품 기획을 세분화하고, 원료·공정·인증·MOQ·투명성 조건을 복잡하게 만드는 방향으로 움직임",
    )
    write_merged(
        card,
        "E7:F8",
        "2026 식품 트렌드에서 기능성, 개인화, AI, 공급망 투명성, 클린라벨, 소량 테스트 니즈가 동시에 확대됨",
    )
    write_merged(
        card,
        "G7:H8",
        "소비자는 건강 목적, 성분 공개, 원산지·인증, 편의성, 가격 부담 완화를 함께 요구함",
    )
    write_merged(
        card,
        "A11:D13",
        "고단백 스낵·음료, 프로바이오틱스 제품, 기능성 RTD, QR 제조정보 공개, 소량 제작·공동구매형 생산",
    )
    write_merged(
        card,
        "E11:H13",
        "만들고 싶은 식품을 입력하면 AI가 제품 유형, 원료, 공정, 인증, MOQ, 지역, 샘플 가능 여부를 분석해 적합한 OEM/ODM 공장을 추천함",
    )
    write_merged(
        card,
        "A16:D18",
        "단백질바, 건강간식, 기능성 음료, 발효식품, 분말/스틱, HMR, 소스/드레싱, 베이커리, 로컬 특산가공",
    )
    write_merged(
        card,
        "E16:H18",
        "발주자가 공장 탐색에서 가장 오래 걸리는 조건은 무엇인가? 공장과 발주자가 제공 가능한 필수 데이터는 무엇인가? 소량 제작 사용자가 먼저 알고 싶은 조건은 무엇인가?",
    )

    for row_idx in [7, 8]:
        card.row_dimensions[row_idx].height = 42
    for row_idx in [11, 12, 13, 16, 17, 18]:
        card.row_dimensions[row_idx].height = 38
    for col in "ABCDEFGH":
        card.column_dimensions[col].width = 18
    card.print_area = "A1:H18"
    card.page_setup.orientation = "landscape"
    card.page_setup.fitToWidth = 1
    card.page_setup.fitToHeight = 1
    card.sheet_properties.pageSetUpPr.fitToPage = True
    card.page_margins.left = 0.25
    card.page_margins.right = 0.25
    card.page_margins.top = 0.35
    card.page_margins.bottom = 0.35

    for ws in [trend, card]:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = copy(cell.alignment)
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal or "left",
                        vertical="center",
                        wrap_text=True,
                    )
                    if ws is trend and 6 <= cell.row <= 15:
                        set_font_size(cell, 8)
                    if ws is card and cell.row in {7, 8, 11, 12, 13, 16, 17, 18}:
                        set_font_size(cell, 9)

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(TARGET)


if __name__ == "__main__":
    main()
